"""
Author: Kelv Gooding
Created: 28/03/2022
Version: 1.005
"""

# Modules

from bs4 import BeautifulSoup
import requests
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Variables

document = "spot-prices.xlsx"
dt = datetime.datetime.today().strftime("%d/%m/%Y")

# Script

def spotprice_atom():
    url = "https://www.coingecko.com/en/coins/cosmos-hub"
    r = requests.get(url)
    soup = BeautifulSoup(r.text, "html.parser")

    # Find all elements, then the attributes relating to this element.

    price = soup.find("span", attrs={"class": "no-wrap"})
    rank = soup.find("div", attrs={"class": "tw-inline-flex tw-items-center tw-px-2 tw-py-0.5 tw-rounded-md tw-text-xs tw-font-medium tw-bg-gray-800 tw-text-gray-100 tw-mb-1 md:tw-mb-0 md:tw-mt-0 dark:tw-bg-gray-600 dark:tw-bg-opacity-40"})

    # Load Document

    wb = load_workbook(document)
    ws = wb["ATOM"]
    column = ws["D"]

    # Print output

    for i in column:
        if i.value == "N/A":
            for item3 in price:
                i.value = item3[1:]
                wb.save(document)
            break
    else:
        for item1, item2 in zip(rank, price):
            ws.append([f"{dt}", f"{item1.text.strip()}", f"{item2.text[1:]}", "N/A"])

    for row in ws[2:ws.max_row]:
        row[0].alignment = Alignment(horizontal='center')
        row[1].alignment = Alignment(horizontal='center')
        row[2].alignment = Alignment(horizontal='center')
        row[3].alignment = Alignment(horizontal='center')
        row[4].alignment = Alignment(horizontal='center')

    wb.save(document)
    wb.close()

def spotprice_algo():
    url = "https://www.coingecko.com/en/coins/algorand"
    r = requests.get(url)
    soup = BeautifulSoup(r.text, "html.parser")

    # Find all elements, then the attributes relating to this element.

    price = soup.find("span", attrs={"class": "no-wrap"})
    rank = soup.find("div", attrs={"class": "tw-inline-flex tw-items-center tw-px-2 tw-py-0.5 tw-rounded-md tw-text-xs tw-font-medium tw-bg-gray-800 tw-text-gray-100 tw-mb-1 md:tw-mb-0 md:tw-mt-0 dark:tw-bg-gray-600 dark:tw-bg-opacity-40"})

    # Load Document

    wb = load_workbook(document)
    ws = wb["ALGO"]
    column = ws["D"]

    # Print output

    for i in column:
        if i.value == "N/A":
            for item3 in price:
                i.value = item3[1:]
                wb.save(document)
            break
    else:
        for item1, item2 in zip(rank, price):
            ws.append([f"{dt}", f"{item1.text.strip()}", f"{item2.text[1:]}", "N/A"])

    for row in ws[2:ws.max_row]:
        row[0].alignment = Alignment(horizontal='center')
        row[1].alignment = Alignment(horizontal='center')
        row[2].alignment = Alignment(horizontal='center')
        row[3].alignment = Alignment(horizontal='center')
        row[4].alignment = Alignment(horizontal='center')

    wb.save(document)
    wb.close()

def spotprice_btc():
    url = "https://www.coingecko.com/en/coins/bitcoin"
    r = requests.get(url)
    soup = BeautifulSoup(r.text, "html.parser")

    # Find all elements, then the attributes relating to this element.

    price = soup.find("span", attrs={"class": "no-wrap"})
    rank = soup.find("div", attrs={"class": "tw-inline-flex tw-items-center tw-px-2 tw-py-0.5 tw-rounded-md tw-text-xs tw-font-medium tw-bg-gray-800 tw-text-gray-100 tw-mb-1 md:tw-mb-0 md:tw-mt-0 dark:tw-bg-gray-600 dark:tw-bg-opacity-40"})

    # Load Document

    wb = load_workbook(document)
    ws = wb["BTC"]
    column = ws["D"]

    # Print output

    for i in column:
        if i.value == "N/A":
            for item3 in price:
                i.value = item3[1:]
                wb.save(document)
            break
    else:
        for item1, item2 in zip(rank, price):
            ws.append([f"{dt}", f"{item1.text.strip()}", f"{item2.text[1:]}", "N/A"])

    for row in ws[2:ws.max_row]:
        row[0].alignment = Alignment(horizontal='center')
        row[1].alignment = Alignment(horizontal='center')
        row[2].alignment = Alignment(horizontal='center')
        row[3].alignment = Alignment(horizontal='center')
        row[4].alignment = Alignment(horizontal='center')

    wb.save(document)
    wb.close()

def spotprice_eth():
    url = "https://www.coingecko.com/en/coins/ethereum"
    r = requests.get(url)
    soup = BeautifulSoup(r.text, "html.parser")

    # Find all elements, then the attributes relating to this element.

    price = soup.find("span", attrs={"class": "no-wrap"})
    rank = soup.find("div", attrs={"class": "tw-inline-flex tw-items-center tw-px-2 tw-py-0.5 tw-rounded-md tw-text-xs tw-font-medium tw-bg-gray-800 tw-text-gray-100 tw-mb-1 md:tw-mb-0 md:tw-mt-0 dark:tw-bg-gray-600 dark:tw-bg-opacity-40"})

    # Load Document

    wb = load_workbook(document)
    ws = wb["ETH"]
    column = ws["D"]

    # Print output

    for i in column:
        if i.value == "N/A":
            for item3 in price:
                i.value = item3[1:]
                wb.save(document)
            break
    else:
        for item1, item2 in zip(rank, price):
            ws.append([f"{dt}", f"{item1.text.strip()}", f"{item2.text[1:]}", "N/A"])

    for row in ws[2:ws.max_row]:
        row[0].alignment = Alignment(horizontal='center')
        row[1].alignment = Alignment(horizontal='center')
        row[2].alignment = Alignment(horizontal='center')
        row[3].alignment = Alignment(horizontal='center')
        row[4].alignment = Alignment(horizontal='center')

    wb.save(document)
    wb.close()

def spotprice_hnt():
    url = "https://www.coingecko.com/en/coins/helium"
    r = requests.get(url)
    soup = BeautifulSoup(r.text, "html.parser")

    # Find all elements, then the attributes relating to this element.

    price = soup.find("span", attrs={"class": "no-wrap"})
    rank = soup.find("div", attrs={"class": "tw-inline-flex tw-items-center tw-px-2 tw-py-0.5 tw-rounded-md tw-text-xs tw-font-medium tw-bg-gray-800 tw-text-gray-100 tw-mb-1 md:tw-mb-0 md:tw-mt-0 dark:tw-bg-gray-600 dark:tw-bg-opacity-40"})

    # Load Document

    wb = load_workbook(document)
    ws = wb["HNT"]
    column = ws["D"]

    # Print output

    for i in column:
        if i.value == "N/A":
            for item3 in price:
                i.value = item3[1:]
                wb.save(document)
            break
    else:
        for item1, item2 in zip(rank, price):
            ws.append([f"{dt}", f"{item1.text.strip()}", f"{item2.text[1:]}", "N/A"])

    for row in ws[2:ws.max_row]:
        row[0].alignment = Alignment(horizontal='center')
        row[1].alignment = Alignment(horizontal='center')
        row[2].alignment = Alignment(horizontal='center')
        row[3].alignment = Alignment(horizontal='center')
        row[4].alignment = Alignment(horizontal='center')

    wb.save(document)
    wb.close()

spotprice_algo()
spotprice_atom()
spotprice_btc()
spotprice_eth()
spotprice_hnt()
